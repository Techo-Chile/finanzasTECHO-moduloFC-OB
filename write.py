# -*- coding: utf-8 -*-
"""Todo:
Agregar control de flujo para ajustar los datos al formato de la nomina"""
from openpyxl import Workbook
from openpyxl.styles import Font, Color, PatternFill, Border, Side
from os.path import join, dirname, abspath
import re

class Writer:
    def __init__(self, writeList=[], writeFormat=[]):
        """Instancia un objeto de clase Writer.

    Args:
      filename (string)       : Fullpath del archivo
      writeList (object[][])  : Lista de filas (cada fila es una lista de objetos)
      writeFormat (int[])     : Lista de enteros en donde cada entero Y en la posición X 
                                indica que el elemento en posición Y de la lista writeList
                                debe ir en la columna X (0-indexada) de la nomina de pago.
                                Ciertos valores deben ser djeados en Nene (Cod Banco), Cuenta"""
        
        '''Formato OB: RUT, Nombre, Modalidad, Código Banco, Cuenta, NFactura, Monto'''
        #self.FORMAT_PROVEEDOR = [0, 1, '3', 16, 17, 2, 3]
        self.FORMAT_PROVEEDOR = [0, 1, '3', 16, 17, 2, 3]
        self.FORMAT_REEMBOLSO = [15, 16, 17, 18, 19, 'rowInd', 4]
        self.FORMAT_REEMBOLSO_FORMATS = ['str', 'str', 'int', 'int', 'str', 'int', 'int']

    def write_reembolso(self,
                        filtered_reemb,
                        outputFileName,
                        personas_list=[None] * 7):
        wb = Workbook()
        ws = wb.active
        self.writeHeader(ws)
        joinedList = self.join_listsByRut(personas_list, filtered_reemb)
        ws.title = 'Hoja1'
        ws = self.get_file_content(ws, joinedList)
        wb.save(outputFileName)

    def join_listsByRut(self, dbFileRows, filteredList):
        ret = []
        for row in filteredList:
            flag = False
            try:
                for rowProveed in dbFileRows:
                    if str(row[2]) in str(rowProveed[0]):
                        ret.append(row + rowProveed)
                        flag = True
                        break
                if not flag:
                    print('RUT {} no encontrado en archivo de personas'.format(row[2]))
            except Exception:
                continue
        return ret
    
    def get_file_content(self, outputFileContent, joinedList):
        newRow = [None] * 29
        for c,row in enumerate(joinedList):
            for rowInd in range(0, len(self.FORMAT_REEMBOLSO)):
                if('rowInd' in str(self.FORMAT_REEMBOLSO[rowInd])):
                    val = (c+1)
                elif(isinstance(self.FORMAT_REEMBOLSO[rowInd], str)):
                    val = self.FORMAT_REEMBOLSO[rowInd]
                else:
                    val = row[self.FORMAT_REEMBOLSO[rowInd]]
                    if '$' in val:
                        val = re.sub("[^0-9]", "", val)
                        row[rowInd] = val
                newRow[rowInd] = str(val) if 'str' in self.FORMAT_REEMBOLSO_FORMATS[rowInd] else int(val)
            newRow[27] = newRow[6]
            outputFileContent.append(newRow)
        return outputFileContent

    def writeHeader(self, ws):
        ws.append([
            'Rut Beneficiario', 'Nombre Beneficiario', 'Cod. Modalidad',
            'Cod Banco', 'Cta Abono', 'N Factura 1', 'Monto 1', 'N Factura 2',
            'Monto 2', 'N Factura 3', 'Monto 3', 'N Factura 4', 'Monto 4',
            'N Factura 5', 'Monto 5', 'N Factura 6', 'Monto 6', 'N Factura 7',
            'Monto 7', 'N Factura 8', 'Monto 8', 'N Factura 9', 'Monto 9',
            'N Factura 10', 'Monto 10', 'N Factura 11', 'Monto 11', 'Monto Total'
        ])
        sd = Side(border_style='thin', color='FF000000')
        for cell in ws["1:1"]:
            cell.font = Font(name='Arial', size=9, color='FF0000FF')
            cell.fill = PatternFill(fill_type='solid', start_color="ffcccccc")
            cell.border = Border(left=sd, right=sd, top=sd, bottom=sd)
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["AC"].width = 15
